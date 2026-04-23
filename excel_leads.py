"""
excel_leads.py
--------------
Captura acumulativa de leads en un Excel local (`leads_pmd.xlsx`).

Cada fila = un lead. Se agrega al final sin sobrescribir.

A futuro este módulo se reemplaza por la integración con PMD System
(el entorno interno del usuario). Mientras tanto, el Excel sirve
como base única de leads para la operación diaria.

Requisitos:
    pip install openpyxl

Uso:
    from excel_leads import append_lead
    append_lead({"origen": "chat", "nombre": "Juan", ...})
"""

from __future__ import annotations

import json
import logging
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import config

logger = logging.getLogger("lucas.excel_leads")

# Directorio JSON donde el dashboard /admin va a leer.
_LEADS_JSON_DIR: Path = config.LOGS_DIR / "leads"
_LEADS_JSON_DIR.mkdir(parents=True, exist_ok=True)


def _save_json_backup(data: dict[str, Any]) -> None:
    """Graba un JSON en logs/leads/ con el payload recibido.
    Así el dashboard /admin (que lee esa carpeta) muestra SIEMPRE todos
    los leads — vengan del chat, del presupuestador, del form de contacto
    o de una reserva de agenda."""
    try:
        ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        origen = (data.get("origen") or "lead").replace("/", "-")
        short_id = uuid.uuid4().hex[:8]
        fname = f"{ts}_{origen}_{short_id}.json"
        (_LEADS_JSON_DIR / fname).write_text(
            json.dumps(data, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
    except Exception as exc:  # pylint: disable=broad-except
        logger.warning("No se pudo grabar JSON backup: %s", exc)

# Archivo acumulativo en logs/ (fuera del repo pero accesible)
EXCEL_PATH: Path = config.LOGS_DIR / "leads_pmd.xlsx"

# Columnas canónicas — el orden se conserva entre filas
COLUMNS: list[str] = [
    "timestamp",
    "origen",            # chat | presupuestador | contacto
    "nombre",
    "email",
    "whatsapp",
    "zona",
    "urgencia",
    "tipo_proyecto",
    "superficie_m2",
    "sistema",
    "nivel",
    "presupuesto_min_usd",
    "presupuesto_max_usd",
    "mensaje",
    "extras",            # libre: JSON-ish
]

# Lock global para evitar condiciones de carrera si llegan 2 leads simultáneos
_write_lock = threading.Lock()


def _ensure_workbook():
    """Crea el workbook con headers si no existe; si existe lo abre."""
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore
    except ImportError:
        logger.warning(
            "openpyxl NO está instalado. El bot sigue funcionando pero los "
            "leads NO se guardan en Excel. Para activarlo: pip install openpyxl"
        )
        return None

    if EXCEL_PATH.exists():
        return load_workbook(EXCEL_PATH)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(COLUMNS)

    # Formato del header — bold + color PMD
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    header_font = Font(bold=True, color="FFFFFF", name="Inter", size=11)
    header_fill = PatternFill("solid", fgColor="2B4E7A")  # --b3
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Anchos razonables
    widths = {
        "A": 22, "B": 16, "C": 22, "D": 28, "E": 20, "F": 20, "G": 24,
        "H": 22, "I": 14, "J": 20, "K": 14, "L": 16, "M": 16, "N": 50, "O": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)
    logger.info("Excel creado en %s", EXCEL_PATH)
    return wb


def append_lead(data: dict[str, Any]) -> bool:
    """Agrega una fila al Excel acumulativo. Thread-safe.

    ADEMÁS guarda un JSON paralelo en logs/leads/ para que el dashboard
    /admin (que lee esa carpeta) vea el lead aunque openpyxl no esté
    instalado o el Excel falle.

    Args:
        data: diccionario con las claves de COLUMNS (las que falten quedan vacías)
              + cualquier extra va a la columna "extras" como texto.

    Returns:
        True si se escribió al Excel, False si falló (el JSON igual se guardó).
    """
    # PRIMERO guardamos el JSON (esto NO puede fallar — el dashboard depende de esto)
    _save_json_backup(data)

    with _write_lock:
        try:
            wb = _ensure_workbook()
            if wb is None:
                # openpyxl no instalado — lead queda solo en logs/leads/ (dispatch_lead)
                return False
            ws = wb["Leads"]

            # Timestamp: si no vino, ahora
            ts = data.get("timestamp") or datetime.now(timezone.utc).isoformat(
                timespec="seconds"
            )

            # Extras: todo lo que no sea columna conocida
            known = set(COLUMNS) | {"timestamp"}
            extras_dict = {k: v for k, v in data.items() if k not in known}
            extras_str = ", ".join(
                f"{k}={v}" for k, v in extras_dict.items() if v not in (None, "", [])
            )

            row = [
                ts,
                data.get("origen", ""),
                data.get("nombre", ""),
                data.get("email", ""),
                data.get("whatsapp") or data.get("phone", ""),
                data.get("zona", ""),
                data.get("urgencia", ""),
                data.get("tipo_proyecto") or data.get("tipo", ""),
                data.get("superficie_m2") or data.get("cubiertos_m2", ""),
                data.get("sistema", ""),
                data.get("nivel", ""),
                data.get("presupuesto_min_usd") or data.get("total_min_usd", ""),
                data.get("presupuesto_max_usd") or data.get("total_max_usd", ""),
                data.get("mensaje") or data.get("comentarios") or data.get("necesidad", ""),
                extras_str,
            ]
            ws.append(row)
            wb.save(EXCEL_PATH)
            logger.info(
                "Lead guardado en Excel (origen=%s email=%s whatsapp=%s)",
                data.get("origen"), data.get("email"), data.get("whatsapp"),
            )
            return True
        except Exception as exc:  # pylint: disable=broad-except
            logger.exception("Error guardando lead en Excel: %s", exc)
            return False


def count_leads() -> int:
    """Cantidad total de leads capturados."""
    try:
        from openpyxl import load_workbook  # type: ignore
        if not EXCEL_PATH.exists():
            return 0
        wb = load_workbook(EXCEL_PATH, read_only=True)
        ws = wb["Leads"]
        return max(ws.max_row - 1, 0)
    except Exception:  # pylint: disable=broad-except
        return 0


def excel_path_str() -> str:
    """Ruta absoluta del Excel (para exponer en health)."""
    return str(EXCEL_PATH.resolve())
